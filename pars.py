import requests
from bs4 import BeautifulSoup as bs
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import fitz

current = {}
#  IUM
url = "https://mccme.ru/ru/nmu/courses-of-nmu/osen-20242025/"
r = requests.get(url)
soup = bs(r.text, "html.parser")
courses = soup.find_all('dd')
for name in courses:
    if name.a != None:
        current[str(name.a)[(str(name.a).index('>') +1):-4]] = name.a['href']

# PDF
pdf_path = '+ course_book_2425.pdf'
raw = []
with fitz.open(pdf_path) as pdf:
    for i in range(1, 6):
        page = pdf[i]
        text = page.get_text()
        if text:
            lines = list(text.split('\n'))
            for _ in lines:
                if len(_) > 5:
                    raw.append(_)
    counter = raw.index("Описания курсов на русском") + 2
    while counter < len(raw):
        if "Course descriptions" in raw[counter]:
            break
        else:
            cc = -1
            if '(' in raw[counter]:
                cc = raw[counter].index("(")
            name = raw[counter][:cc]
            if raw[counter][-1].isnumeric():
                page = raw[counter][-3::]
            else:
                page = raw[counter+1][-3::]
                counter+=1
            counter +=1
            current[name] = page

df = pd.DataFrame(list(current.items()), columns=['name', 'link'])
df['lvl'] = ''
df['algem'] = ''
df['Calculus'] = ''
df['Geometry'] = ''
df['Algebra'] = ''
df['Applied'] = ''
df['Physics'] = ''
df['Number Theory'] = ''
df['Topology'] = ''
df['Logic'] = ''
df['Math.Stat and Probability'] = ''
df['Economics'] = ''
df['CS and DS'] = ''
df['Diff Gem'] = ''
df['Dynamics and Diff.Equations'] = ''

df.to_excel('Raw_Table.xlsx', index=False)
wb = load_workbook('Raw_Table.xlsx')
ws = wb.active
ws.column_dimensions['A'].width = 60  # Ширина первого столбца (Ключи)
ws.column_dimensions['B'].width = 20  # Ширина второго столбца (Значения)
for col in range(3, 27):  # 1 до 26 соответствует A-Z
    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

wb.save('Raw_Table.xlsx')


